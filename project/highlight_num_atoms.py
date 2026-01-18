from rdkit import Chem
from rdkit.Chem import Draw
from rdkit.Chem.Draw import rdMolDraw2D
from rdkit import Chem
from rdkit.Chem import Draw
from PIL import Image
import io


# def highlight_num_atoms(smiles, atom_indices):
#     """
#     根据给定的原子索引高亮 SMILES 表示的分子中的原子。
#     :param smiles: 分子的 SMILES 表示
#     :param atom_indices: 需要高亮的原子索引列表
#     """
#     # 使用 RDKit 从 SMILES 创建分子对象
#     mol = Chem.MolFromSmiles(smiles)
#     if not mol:
#         raise ValueError("Invalid SMILES string")
#
#     # highlight_colors={0:(0,0,255),1:(0,0,255)}
#
#     # 创建分子绘图对象，并高亮指定的原子
#     drawer = Draw.MolDraw2DCairo(300, 300)
#     opts = drawer.drawOptions()
#     drawer.DrawMolecule(mol, highlightAtoms=atom_indices)
#     # drawer.DrawMolecule(mol, highlightAtoms=atom_indices)
#     drawer.FinishDrawing()
#     img = drawer.GetDrawingText()
#     return img
#
# if __name__ == "__main__":
#     # highlight atom
#     #smiles="ClC1=CC=CC=C1I.C1(NC2=CC=CC=C2)=CC=CC=C1"
#     smiles="C=CCC(CC#CCC(CC#CC)C(OCC)=O)(C(OCC)=O)C(OCC)=O"
#     location=[15, 16, 20, 21]
#     atom_indices=location
#     img_data=highlight_num_atoms(smiles, atom_indices)
#     img = Image.open(io.BytesIO(img_data))
#     img.show()

import io
import os
import ast

import pandas as pd
from rdkit import Chem
from rdkit.Chem import Draw
from PIL import Image


# def highlight_num_atoms(smiles, atom_indices):
#     """
#     根据给定的原子索引高亮 SMILES 表示的分子中的原子。
#     :param smiles: 分子的 SMILES 表示
#     :param atom_indices: 需要高亮的原子索引列表 (基于 RDKit atom index)
#     :return: PNG 图片的二进制数据
#     """
#     # 使用 RDKit 从 SMILES 创建分子对象
#     mol = Chem.MolFromSmiles(smiles)
#     if mol is None:
#         raise ValueError(f"Invalid SMILES string: {smiles}")
#
#     # 创建分子绘图对象，并高亮指定的原子
#     drawer = Draw.MolDraw2DCairo(300, 300)
#     drawer.DrawMolecule(mol, highlightAtoms=atom_indices)
#     drawer.FinishDrawing()
#     img_bytes = drawer.GetDrawingText()  # PNG 的二进制数据
#     return img_bytes
#
#
# def smiles_excel_to_images(
#     excel_in,
#     excel_out,
#     img_dir="atom_images",
#     smiles_col="smiles",
#     atom_idx_col="atom_indices",
#     img_col="atom_img",
# ):
#     """
#     读取一个 Excel，使用 highlight_num_atoms 生成高亮图片，
#     将图片保存到本地，并把图片路径写入到 Excel 的 img_col 列。
#
#     :param excel_in: 输入 Excel 路径
#     :param excel_out: 输出 Excel 路径
#     :param img_dir: 保存图片的文件夹
#     :param smiles_col: SMILES 所在列名
#     :param atom_idx_col: 原子索引所在列名（如 "[1, 2, 3]" 这种字符串）
#     :param img_col: 要写入图片路径的列名
#     """
#     os.makedirs(img_dir, exist_ok=True)
#
#     df = pd.read_excel(excel_in)
#
#     img_paths = []
#
#     for idx, row in df.iterrows():
#         smiles = row.get(smiles_col, None)
#         atom_indices_raw = row.get(atom_idx_col, None)
#
#         # 遇到缺失值就跳过
#         if pd.isna(smiles) or pd.isna(atom_indices_raw):
#             img_paths.append(None)
#             continue
#
#         # 解析 atom_indices：一般是字符串 "[15, 16, 20, 21]" 这种
#         try:
#             if isinstance(atom_indices_raw, str):
#                 atom_indices = ast.literal_eval(atom_indices_raw)
#             elif isinstance(atom_indices_raw, (list, tuple)):
#                 atom_indices = list(atom_indices_raw)
#             else:
#                 # 单个数字的情况
#                 atom_indices = [int(atom_indices_raw)]
#         except Exception as e:
#             print(f"Row {idx}: parse atom_indices error: {atom_indices_raw} ({e})")
#             img_paths.append(None)
#             continue
#
#         try:
#             img_bytes = highlight_num_atoms(smiles, atom_indices)
#             img = Image.open(io.BytesIO(img_bytes))
#
#             img_path = os.path.join(img_dir, f"row_{idx}.png")
#             img.save(img_path)
#             img_paths.append(img_path)
#         except Exception as e:
#             print(f"Row {idx}: draw/save error: {e}")
#             img_paths.append(None)
#
#     # 把图片路径写入新列
#     df[img_col] = img_paths
#     df.to_excel(excel_out, index=False)
#     print(f"Done. Saved result to {excel_out}, images in {img_dir}/")
#
#
# if __name__ == "__main__":
#     # # 单例测试
#     # smiles = "C=CCC(CC#CCC(CC#CC)C(OCC)=O)(C(OCC)=O)C(OCC)=O"
#     # location = [15, 16, 20, 21]
#     # img_data = highlight_num_atoms(smiles, location)
#     # img = Image.open(io.BytesIO(img_data))
#     # img.show()
#
#     # 批量处理 Excel
#     # 假设你的 Excel 有两列： "smiles" 和 "atom_indices"
#     # atom_indices 列的格式类似 "[15, 16, 20, 21]"
#     excel_in = "network_transformer_50k_mcc_0.71_atoms_index.xlsx"
#     excel_out = "output_with_images.xlsx"
#
#     smiles_excel_to_images(
#         excel_in=excel_in,
#         excel_out=excel_out,
#         img_dir="atom_images",
#         smiles_col="smiles",
#         atom_idx_col="atom_indices",
#         img_col="atom_img",
#     )

# import os
# import ast
#
# import pandas as pd
# from rdkit import Chem
# from rdkit.Chem import Draw
#
#
# def highlight_num_atoms(smiles, atom_indices, size=(300, 300)):
#     """
#     根据给定的原子索引高亮 SMILES 表示的分子中的原子。
#     :param smiles: 分子的 SMILES 字符串
#     :param atom_indices: 需要高亮的原子索引列表 (基于 RDKit atom index)
#     :param size: 图片尺寸
#     :return: PIL.Image 对象
#     """
#     mol = Chem.MolFromSmiles(smiles)
#     if mol is None:
#         raise ValueError(f"Invalid SMILES string: {smiles}")
#
#     # 直接用 MolToImage 画图并高亮
#     img = Draw.MolToImage(mol, size=size, highlightAtoms=atom_indices)
#     return img
#
#
# def excel_to_highlight_images(
#     excel_in,
#     excel_out,
#     img_dir="atom_images",
#     smiles_col="smiles",
#     atom_idx_col="atom_index",
#     img_col="atom_img",
# ):
#     """
#     读取 Excel，使用 highlight_num_atoms 生成图片，
#     将图片保存到本地，并把图片路径写入到 img_col 列。
#
#     :param excel_in: 输入 Excel 路径
#     :param excel_out: 输出 Excel 路径（注意：结果保存在这个文件里）
#     :param img_dir: 保存图片的文件夹
#     :param smiles_col: SMILES 所在列名
#     :param atom_idx_col: 原子索引列表所在列名（如 "[14, 17, 19, 20]"）
#     :param img_col: 要写入图片路径的列名
#     """
#     os.makedirs(img_dir, exist_ok=True)
#
#     df = pd.read_excel(excel_in)
#
#     img_paths = []
#
#     for idx, row in df.iterrows():
#         smiles = row.get(smiles_col, None)
#         atom_indices_raw = row.get(atom_idx_col, None)
#
#         # 遇到缺失值就跳过
#         if pd.isna(smiles):
#             print(f"Row {idx}: smiles is NaN, skip")
#             img_paths.append(None)
#             continue
#
#         # 解析 atom_index 列
#         if pd.isna(atom_indices_raw):
#             # 原子索引为空时，可以画一个不高亮的分子
#             atom_indices = []
#         else:
#             try:
#                 if isinstance(atom_indices_raw, str):
#                     # 例如 "[14, 17, 19, 20]"
#                     atom_indices = ast.literal_eval(atom_indices_raw)
#                 elif isinstance(atom_indices_raw, (list, tuple)):
#                     atom_indices = list(atom_indices_raw)
#                 else:
#                     # 单个数字
#                     atom_indices = [int(atom_indices_raw)]
#             except Exception as e:
#                 print(f"Row {idx}: parse atom_index error: {atom_indices_raw} ({e})")
#                 img_paths.append(None)
#                 continue
#
#         try:
#             img = highlight_num_atoms(smiles, atom_indices)
#             img_path = os.path.join(img_dir, f"row_{idx}.png")
#             img.save(img_path)
#             img_paths.append(img_path)
#
#             if idx < 5:  # 前几行打印一下日志
#                 print(f"Row {idx}: saved {img_path}, indices={atom_indices}")
#         except Exception as e:
#             print(f"Row {idx}: draw/save error: {e}")
#             img_paths.append(None)
#
#     # 把图片路径写入新列
#     df[img_col] = img_paths
#     df.to_excel(excel_out, index=False)
#     print(f"\n处理完成：结果保存在 {excel_out}")
#     print(f"图片保存在目录：{img_dir}/")
#
#
# if __name__ == "__main__":
#     # ====== 单个测试（可选） ======
#     smiles = "C=CCC(CC#CCC(CC#CC)C(OCC)=O)(C(OCC)=O)C(OCC)=O"
#     location = [15, 16, 20, 21]
#     img = highlight_num_atoms(smiles, location)
#     img.show()  # 看看单个分子是否正常显示
#
#     # ====== 批量处理 Excel ======
#     excel_in = "network_transformer_50k_mcc_0.71_atoms_index.xlsx"         # ← 把这里改成你那张表的路径
#     excel_out = "with_atom_img.xlsx"     # ← 输出的新文件名
#
#     excel_to_highlight_images(
#         excel_in=excel_in,
#         excel_out=excel_out,
#         img_dir="atom_images",
#         smiles_col="smiles",      # B 列列名
#         atom_idx_col="atom_index",  # E 列列名
#         img_col="atom_img",       # F 列（输出）
#     )


# import os
# import ast
#
# import pandas as pd
# from rdkit import Chem
# from rdkit.Chem import Draw
#
# from openpyxl import Workbook, load_workbook
# from openpyxl.drawing.image import Image as XLImage
# from openpyxl.utils import get_column_letter
#
#
# def highlight_num_atoms(smiles, atom_indices, size=(300, 300)):
#     """
#     根据给定的原子索引高亮 SMILES 表示的分子中的原子。
#     :param smiles: SMILES 字符串
#     :param atom_indices: 需要高亮的原子索引列表
#     :param size: 图片尺寸
#     :return: PIL.Image 对象
#     """
#     mol = Chem.MolFromSmiles(smiles)
#     if mol is None:
#         raise ValueError(f"Invalid SMILES string: {smiles}")
#     img = Draw.MolToImage(mol, size=size, highlightAtoms=atom_indices)
#     return img
#
#
# def excel_embed_highlight_images(
#     excel_in,
#     excel_out,
#     img_dir="atom_images",
#     smiles_col="smiles",
#     atom_idx_col="atom_index",
#     img_col="atom_img",
# ):
#     """
#     读取 Excel，生成高亮图片，并把图片嵌入到 excel_out 的 img_col 列单元格中。
#
#     :param excel_in: 输入 Excel 文件路径
#     :param excel_out: 输出 Excel 文件路径
#     :param img_dir: 保存 PNG 图片的文件夹
#     :param smiles_col: SMILES 列名
#     :param atom_idx_col: 原子索引列名（如 "[14, 17, 19, 20]"）
#     :param img_col: 要插入图片的列名
#     """
#     os.makedirs(img_dir, exist_ok=True)
#
#     # 1) 用 pandas 读原始表
#     df = pd.read_excel(excel_in)
#
#     # 2) 先把原表写到新的 Excel（方便用 openpyxl 操作）
#     df.to_excel(excel_out, index=False)
#
#     # 3) 加载刚写出的 Excel，用 openpyxl 嵌入图片
#     wb = load_workbook(excel_out)
#     ws = wb.active
#
#     # 找到图片列的列号（1-based）
#     cols = list(df.columns)
#     if img_col not in cols:
#         # 如果原表没有 atom_img 列，就追加一列
#         cols.append(img_col)
#         df[img_col] = ""
#         # 重新写一次并重新加载
#         df.to_excel(excel_out, index=False)
#         wb = load_workbook(excel_out)
#         ws = wb.active
#     img_col_idx = cols.index(img_col) + 1
#     img_col_letter = get_column_letter(img_col_idx)
#
#     # 调整图片所在列的宽度
#     ws.column_dimensions[img_col_letter].width = 25
#
#     # 4) 一行一行生成图片并插入
#     for idx, row in df.iterrows():
#         smiles = row.get(smiles_col, None)
#         atom_indices_raw = row.get(atom_idx_col, None)
#
#         if pd.isna(smiles):
#             continue
#
#         # 解析 atom_index 列
#         if pd.isna(atom_indices_raw):
#             atom_indices = []
#         else:
#             try:
#                 if isinstance(atom_indices_raw, str):
#                     atom_indices = ast.literal_eval(atom_indices_raw)
#                 elif isinstance(atom_indices_raw, (list, tuple)):
#                     atom_indices = list(atom_indices_raw)
#                 else:
#                     atom_indices = [int(atom_indices_raw)]
#             except Exception as e:
#                 print(f"Row {idx}: parse atom_index error: {atom_indices_raw} ({e})")
#                 continue
#
#         try:
#             img = highlight_num_atoms(smiles, atom_indices)
#             img_path = os.path.join(img_dir, f"row_{idx}.png")
#             img.save(img_path)
#
#             xl_img = XLImage(img_path)
#
#             # 目标单元格（注意 Excel 从第 2 行开始是数据）
#             excel_row = idx + 2
#             cell = ws.cell(row=excel_row, column=img_col_idx)
#             xl_img.anchor = cell.coordinate
#             ws.add_image(xl_img)
#
#             # 调整这一行高度，让图片能完整显示
#             ws.row_dimensions[excel_row].height = 80
#
#             if idx < 5:
#                 print(f"Row {idx}: insert image {img_path}, indices={atom_indices}")
#         except Exception as e:
#             print(f"Row {idx}: draw/save/insert error: {e}")
#             continue
#
#     wb.save(excel_out)
#     print(f"\n完成：图片已嵌入到 {excel_out} 的 '{img_col}' 列。")
#     print(f"PNG 图片保存在目录：{img_dir}/")
#
#
# if __name__ == "__main__":
#     # 只需要改这两个路径
#     excel_in = "network_transformer_50k_mcc_0.71_atoms_index.xlsx"       # 原始那张表（有 smiles, atom_index, atom_img 列）
#     excel_out = "with_atom_img_1.xlsx"   # 输出的新文件
#
#     excel_embed_highlight_images(
#         excel_in=excel_in,
#         excel_out=excel_out,
#         img_dir="atom_images",
#         smiles_col="smiles",        # 和你表里列名对应
#         atom_idx_col="atom_index",  # E 列
#         img_col="atom_img",         # F 列
#     )

# import os
# import ast
#
# import pandas as pd
# from rdkit import Chem
# from rdkit.Chem import Draw
#
# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image as XLImage
# from openpyxl.utils import get_column_letter
#
#
# def highlight_num_atoms(smiles, atom_indices, size=(120, 120)):
#     """
#     根据给定的原子索引高亮 SMILES 分子中的原子和相邻键。
#     - 图片尺寸缩小为 size（默认 120x120）
#     - 如果一条键的两个原子都在 atom_indices 中，则整条键高亮
#     """
#     mol = Chem.MolFromSmiles(smiles)
#     if mol is None:
#         raise ValueError(f"Invalid SMILES string: {smiles}")
#
#     # 计算需要高亮的 bond：两端原子都在 atom_indices 中
#     atom_set = set(atom_indices)
#     bond_indices = []
#     for bond in mol.GetBonds():
#         a1 = bond.GetBeginAtomIdx()
#         a2 = bond.GetEndAtomIdx()
#         if a1 in atom_set and a2 in atom_set:
#             bond_indices.append(bond.GetIdx())
#
#     # 用 MolToImage 画图并高亮原子和键（图片更小）
#     img = Draw.MolToImage(
#         mol,
#         size=size,
#         highlightAtoms=atom_indices,
#         highlightBonds=bond_indices,
#     )
#     return img
#
#
# def excel_embed_highlight_images(
#     excel_in,
#     excel_out,
#     img_dir="atom_images",
#     smiles_col="smiles",
#     atom_idx_col="atom_index",
#     img_col="atom_img",
# ):
#     """
#     读取 Excel，生成高亮图片，并把图片嵌入到 excel_out 的 img_col 列单元格中。
#     """
#     os.makedirs(img_dir, exist_ok=True)
#
#     # 用 pandas 读原始表
#     df = pd.read_excel(excel_in)
#
#     # 先把原表写到新的 Excel
#     df.to_excel(excel_out, index=False)
#
#     # 用 openpyxl 打开
#     wb = load_workbook(excel_out)
#     ws = wb.active
#
#     # 确定图片列索引
#     cols = list(df.columns)
#     if img_col not in cols:
#         cols.append(img_col)
#         df[img_col] = ""
#         df.to_excel(excel_out, index=False)
#         wb = load_workbook(excel_out)
#         ws = wb.active
#
#     img_col_idx = cols.index(img_col) + 1
#     img_col_letter = get_column_letter(img_col_idx)
#
#     # 调小列宽以适配小图片
#     ws.column_dimensions[img_col_letter].width = 18
#
#     for idx, row in df.iterrows():
#         smiles = row.get(smiles_col, None)
#         atom_indices_raw = row.get(atom_idx_col, None)
#
#         if pd.isna(smiles):
#             continue
#
#         # 解析 atom_index 列
#         if pd.isna(atom_indices_raw):
#             atom_indices = []
#         else:
#             try:
#                 if isinstance(atom_indices_raw, str):
#                     atom_indices = ast.literal_eval(atom_indices_raw)
#                 elif isinstance(atom_indices_raw, (list, tuple)):
#                     atom_indices = list(atom_indices_raw)
#                 else:
#                     atom_indices = [int(atom_indices_raw)]
#             except Exception as e:
#                 print(f"Row {idx}: parse atom_index error: {atom_indices_raw} ({e})")
#                 continue
#
#         try:
#             # 这里用缩小后的 size
#             img = highlight_num_atoms(smiles, atom_indices, size=(120, 120))
#             img_path = os.path.join(img_dir, f"row_{idx}.png")
#             img.save(img_path)
#
#             xl_img = XLImage(img_path)
#
#             # 目标单元格（Excel 数据从第 2 行开始）
#             excel_row = idx + 2
#             cell = ws.cell(row=excel_row, column=img_col_idx)
#             xl_img.anchor = cell.coordinate
#             ws.add_image(xl_img)
#
#             # 调整行高（比之前小一点）
#             ws.row_dimensions[excel_row].height = 65
#
#             if idx < 5:
#                 print(f"Row {idx}: inserted image {img_path}, atoms={atom_indices}")
#         except Exception as e:
#             print(f"Row {idx}: draw/save/insert error: {e}")
#             continue
#
#     wb.save(excel_out)
#     print(f"\n完成：图片已嵌入到 {excel_out} 的 '{img_col}' 列。")
#     print(f"PNG 图片保存在目录：{img_dir}/")
#
#
# if __name__ == "__main__":
#     # 按你的实际路径修改
#     excel_in = "network_transformer_50k_mcc_0.71_atoms_index.xlsx"
#     excel_out = "with_atom_img_small_bond.xlsx"
#
#     excel_embed_highlight_images(
#         excel_in=excel_in,
#         excel_out=excel_out,
#         img_dir="atom_images_small",
#         smiles_col="smiles",
#         atom_idx_col="atom_index",
#         img_col="atom_img",
#     )

import os
import ast

import pandas as pd
from rdkit import Chem
from rdkit.Chem import Draw

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


def highlight_num_atoms(smiles, atom_indices, size=(120, 120), internal_size=(400, 400)):
    """
    根据给定的原子索引高亮 SMILES 分子中的原子和相邻键，并返回一个小而清晰的图像。

    - 先用 internal_size 较大尺寸绘制，再用 LANCZOS 高质量缩放到 size。
    - 如果一条键的两个原子都在 atom_indices 中，则整条键一起高亮。
    """
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        raise ValueError(f"Invalid SMILES string: {smiles}")

    atom_set = set(atom_indices)
    bond_indices = []
    for bond in mol.GetBonds():
        a1 = bond.GetBeginAtomIdx()
        a2 = bond.GetEndAtomIdx()
        if a1 in atom_set and a2 in atom_set:
            bond_indices.append(bond.GetIdx())

    # 先用较大尺寸绘图
    big_img = Draw.MolToImage(
        mol,
        size=internal_size,
        highlightAtoms=atom_indices,
        highlightBonds=bond_indices,
    )

    # 再缩放到目标尺寸，用高质量插值
    if big_img.size != size:
        big_img = big_img.resize(size, resample=PILImage.LANCZOS)

    return big_img


def excel_embed_highlight_images(
    excel_in,
    excel_out,
    img_dir="atom_images",
    smiles_col="smiles",
    atom_idx_col="atom_index",
    img_col="atom_img",
    img_size=(120, 120),
):
    """
    读取 Excel，生成高亮图片，并把图片嵌入到 excel_out 的 img_col 列单元格中。
    """
    os.makedirs(img_dir, exist_ok=True)

    df = pd.read_excel(excel_in)
    df.to_excel(excel_out, index=False)

    wb = load_workbook(excel_out)
    ws = wb.active

    cols = list(df.columns)
    if img_col not in cols:
        cols.append(img_col)
        df[img_col] = ""
        df.to_excel(excel_out, index=False)
        wb = load_workbook(excel_out)
        ws = wb.active

    img_col_idx = cols.index(img_col) + 1
    img_col_letter = get_column_letter(img_col_idx)

    # 列宽、行高大一点，尽量避免 Excel 再缩放图像
    ws.column_dimensions[img_col_letter].width = 20

    for idx, row in df.iterrows():
        smiles = row.get(smiles_col, None)
        atom_indices_raw = row.get(atom_idx_col, None)

        if pd.isna(smiles):
            continue

        if pd.isna(atom_indices_raw):
            atom_indices = []
        else:
            try:
                if isinstance(atom_indices_raw, str):
                    atom_indices = ast.literal_eval(atom_indices_raw)
                elif isinstance(atom_indices_raw, (list, tuple)):
                    atom_indices = list(atom_indices_raw)
                else:
                    atom_indices = [int(atom_indices_raw)]
            except Exception as e:
                print(f"Row {idx}: parse atom_index error: {atom_indices_raw} ({e})")
                continue

        try:
            img = highlight_num_atoms(
                smiles,
                atom_indices,
                size=img_size,
                internal_size=(400, 400),  # 内部画大图
            )
            img_path = os.path.join(img_dir, f"row_{idx}.png")
            img.save(img_path)

            xl_img = XLImage(img_path)
            # 保证插入到 Excel 时就是我们想要的大小（单位：像素）
            xl_img.width, xl_img.height = img_size

            excel_row = idx + 2
            cell = ws.cell(row=excel_row, column=img_col_idx)
            xl_img.anchor = cell.coordinate
            ws.add_image(xl_img)

            # 行高用点（pt），粗略按 0.75 * 像素
            ws.row_dimensions[excel_row].height = img_size[1] * 0.75

            if idx < 5:
                print(f"Row {idx}: inserted image {img_path}, atoms={atom_indices}")
        except Exception as e:
            print(f"Row {idx}: draw/save/insert error: {e}")
            continue

    wb.save(excel_out)
    print(f"\n完成：图片已嵌入到 {excel_out} 的 '{img_col}' 列。")
    print(f"PNG 图片保存在目录：{img_dir}/")


if __name__ == "__main__":
    excel_in = "network_transformer_7_7_50k_0.71_atoms_ruizhen_product_test_dot_index.xlsx"
    excel_out = "network_transformer_7_7_50k_0.71_atoms_ruizhen_product_test_dot_index_img.xlsx"

    excel_embed_highlight_images(
        excel_in=excel_in,
        excel_out=excel_out,
        img_dir="atom_images_clear",
        smiles_col="smiles",
        atom_idx_col="atom_index",
        img_col="atom_img",
        img_size=(120, 120),  # 可以改成 (100, 100) / (140, 140) 微调
    )

