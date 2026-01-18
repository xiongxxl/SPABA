import pandas as pd
from rdkit import Chem
from rdkit.Chem import Draw
from rdkit.Chem.Draw import rdMolDraw2D
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import os
import io

import pandas as pd
from rdkit import Chem
from rdkit.Chem import Draw
from rdkit.Chem.Draw import rdMolDraw2D
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import os
import io
import ast


def highlight_molecule(smiles, atom_indices, size=(300, 300)):
    """生成高亮指定原子的分子图像"""
    mol = Chem.MolFromSmiles(smiles)
    if not mol:
        return None

    # 转换原子索引为整数列表
    try:
        if isinstance(atom_indices, str):
            atom_indices = ast.literal_eval(atom_indices)
        highlight_atoms = [int(i) for i in atom_indices]
    except:
        return None

    # 设置绘图选项
    drawer = rdMolDraw2D.MolDraw2DCairo(size[0], size[1])
    opts = drawer.drawOptions()
    opts.addAtomIndices = True
    opts.highlightRadius = 0.3

    # 设置高亮颜色（金色）
    highlight_color = (1, 0.84, 0, 0.5)  # RGBA格式的金色

    # 绘制高亮分子
    drawer.DrawMolecule(mol,
                        highlightAtoms=highlight_atoms,
                        highlightAtomColors={i: highlight_color for i in highlight_atoms})
    drawer.FinishDrawing()

    img = Image.open(io.BytesIO(drawer.GetDrawingText()))
    return img


def process_excel_with_highlight(input_path, output_path):
    """处理Excel文件并添加高亮图像"""
    try:
        # 读取Excel文件
        df = pd.read_excel(input_path)
        print(f"成功读取 {len(df)} 行数据")

        # 检查必要列
        required_cols = ['smiles', 'gold_criterion']
        for col in required_cols:
            if col not in df.columns:
                raise ValueError(f"缺少必要列: {col}")

        # 创建临时目录
        temp_dir = 'temp_highlight_imgs'
        os.makedirs(temp_dir, exist_ok=True)

        # 使用openpyxl加载原始Excel以保留格式
        wb = load_workbook(input_path)
        ws = wb.active

        # 确定criterion_img列位置
        if 'criterion_img' not in df.columns:
            img_col = len(df.columns) + 1
            ws.cell(row=1, column=img_col, value='criterion_img')
        else:
            img_col = df.columns.get_loc('criterion_img') + 1

        # 处理每一行
        for idx, row in df.iterrows():
            smiles = str(row['smiles']) if pd.notna(row['smiles']) else None
            atom_indices = row['gold_criterion']

            if not smiles or pd.isna(atom_indices):
                continue

            try:
                # 生成高亮图像
                img = highlight_molecule(smiles, atom_indices)
                if not img:
                    print(f"行 {idx + 2}: 无法处理 - SMILES: {smiles}, 原子索引: {atom_indices}")
                    continue

                # 保存临时图像
                img_path = f"{temp_dir}/highlight_{idx}.png"
                img.save(img_path, dpi=(300, 300))

                # 添加到Excel
                excel_img = ExcelImage(img_path)
                excel_img.width, excel_img.height = 200, 200
                ws.add_image(excel_img, f"{chr(64 + img_col)}{idx + 2}")

                print(f"行 {idx + 2}: 成功添加高亮图像")

            except Exception as e:
                print(f"行 {idx + 2} 处理错误: {str(e)}")

        # 调整图像列宽度
        ws.column_dimensions[chr(64 + img_col)].width = 30

        # 保存结果
        wb.save(output_path)
        print(f"\n✅ 处理完成！结果保存到: {output_path}")

    finally:
        # 清理临时文件
        if os.path.exists(temp_dir):
            for f in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, f))
            os.rmdir(temp_dir)


if __name__ == "__main__":
    input_file = "multi_criterion_1500_orign_del_large.xlsx"
    output_file = "output_with_highlight.xlsx"

    if not os.path.exists(input_file):
        print(f"❌ 错误: 输入文件不存在 {input_file}")
    else:
        process_excel_with_highlight(input_file, output_file)





# def draw_molecule_with_atom_numbers(mol, size=(200, 200)):
#     """使用rdMolDraw2D绘制带原子编号的分子"""
#     drawer = rdMolDraw2D.MolDraw2DCairo(size[0], size[1])
#
#     # 关键设置：启用原子编号显示
#     drawer.drawOptions().addAtomIndices = True
#     drawer.drawOptions().atomLabelDeuteriumTritium = True
#     drawer.drawOptions().highlightRadius = 0.3
#
#     # 绘制分子
#     drawer.DrawMolecule(mol)
#     drawer.FinishDrawing()
#
#     # 转换为PIL图像
#     img = Image.open(io.BytesIO(drawer.GetDrawingText()))
#     return img
#
#
# def smiles_to_image(smiles, size=(200, 200)):
#     """从SMILES生成带原子编号的图像"""
#     mol = Chem.MolFromSmiles(smiles)
#     if not mol:
#         return None
#
#     # 添加显式原子编号（从0开始）
#     for atom in mol.GetAtoms():
#         atom.SetProp("atomNote", str(atom.GetIdx()))
#
#     return draw_molecule_with_atom_numbers(mol, size)
#
#
# def process_excel(input_path, output_path):
#     """处理Excel文件主函数"""
#     try:
#         # 读取Excel文件
#         df = pd.read_excel(input_path)
#
#         # 检查必要列
#         if 'smiles' not in df.columns:
#             raise ValueError("Excel中必须包含'smiles'列")
#
#         # 创建新工作簿
#         wb = Workbook()
#         ws = wb.active
#         ws.append(list(df.columns) + ['code_img'])  # 添加标题行
#
#         # 创建临时目录
#         temp_dir = 'temp_mol_imgs'
#         os.makedirs(temp_dir, exist_ok=True)
#
#         # 处理每一行
#         for idx, row in df.iterrows():
#             smiles = row['smiles']
#             if pd.isna(smiles):
#                 continue
#
#             try:
#                 # 生成分子图像
#                 img = smiles_to_image(str(smiles))
#                 if not img:
#                     print(f"⚠️ 无法处理SMILES: {smiles} (行 {idx + 2})")
#                     continue
#
#                 # 保存临时图像
#                 img_path = f"{temp_dir}/mol_{idx}.png"
#                 img.save(img_path)
#
#                 # 添加到Excel
#                 excel_img = ExcelImage(img_path)
#                 excel_img.width, excel_img.height = 200, 200
#                 img_col = len(df.columns) + 1
#                 ws.add_image(excel_img, f"{chr(64 + img_col)}{idx + 2}")
#
#                 # 写入其他数据
#                 for col_idx, value in enumerate(row, 1):
#                     ws.cell(row=idx + 2, column=col_idx, value=value)
#
#             except Exception as e:
#                 print(f"处理行 {idx + 2} 出错: {e}")
#
#         # 调整单元格大小
#         img_col_letter = chr(65 + len(df.columns))
#         ws.column_dimensions[img_col_letter].width = 25
#         for row in range(2, len(df) + 2):
#             ws.row_dimensions[row].height = 150
#
#         # 保存结果
#         wb.save(output_path)
#         print(f"✅ 处理完成，结果保存到: {output_path}")
#
#     finally:
#         # 清理临时文件
#         if os.path.exists(temp_dir):
#             for f in os.listdir(temp_dir):
#                 os.remove(os.path.join(temp_dir, f))
#             os.rmdir(temp_dir)
#
#
# if __name__ == "__main__":
#     current_dir = os.getcwd()
#     parent_dir = os.path.dirname(current_dir)
#     label_file_name = f'data/result/statistics_supervision'
#     label_excel = os.path.join(parent_dir, label_file_name)
#     input_file=os.path.join(label_excel, 'multi_criterion_1500_orign_v02.xlsx')
#     output_file=os.path.join(label_excel, 'multi_criterion_1500_orign_img_v02.xlsx')
#     # input_file = "multi_criterion_1500_orign_v02.xlsx"  # 替换为你的输入文件
#     # output_file = "output_with_atom_numbers.xlsx"  # 输出文件名
#
#     if not os.path.exists(input_file):
#         print(f"❌ 错误: 输入文件不存在 {input_file}")
#     else:
#         process_excel(input_file, output_file)
