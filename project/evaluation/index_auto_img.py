import os
import ast
import pandas as pd
from rdkit import Chem
from rdkit.Chem import Draw
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


def highlight_num_atoms(smiles, atom_indices, size=(120, 120), internal_size=(400, 400)):
    """
    根据给定的原子索引高亮 SMILES 分子中的原子和相邻键，并返回一个小而清晰的图像。

    - 先用 internal_size 较大尺寸绘制，再用 LANCZOS 高质量缩放到 size。
    - 如果一条键的两个原子都在 atom_indices 中，则整条键一起高亮。
    """
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        raise ValueError(f"Invalid SMILES string: {smiles}")

    atom_set = set(atom_indices)
    bond_indices = []
    for bond in mol.GetBonds():
        a1 = bond.GetBeginAtomIdx()
        a2 = bond.GetEndAtomIdx()
        if a1 in atom_set and a2 in atom_set:
            bond_indices.append(bond.GetIdx())

    # 先用较大尺寸绘图
    big_img = Draw.MolToImage(
        mol,
        size=internal_size,
        highlightAtoms=atom_indices,
        highlightBonds=bond_indices,
    )

    # 再缩放到目标尺寸，用高质量插值
    if big_img.size != size:
        big_img = big_img.resize(size, resample=PILImage.LANCZOS)

    return big_img


def excel_embed_highlight_images(
    excel_in,
    excel_out,
    img_dir="atom_images",
    smiles_col="smiles",
    atom_idx_col="atom_index",
    img_col="atom_img",
    img_size=(120, 120),
):
    """
    读取 Excel，生成高亮图片，并把图片嵌入到 excel_out 的 img_col 列单元格中。
    """
    os.makedirs(img_dir, exist_ok=True)

    df = pd.read_excel(excel_in)
    df.to_excel(excel_out, index=False)

    wb = load_workbook(excel_out)
    ws = wb.active

    cols = list(df.columns)
    if img_col not in cols:
        cols.append(img_col)
        df[img_col] = ""
        df.to_excel(excel_out, index=False)
        wb = load_workbook(excel_out)
        ws = wb.active

    img_col_idx = cols.index(img_col) + 1
    img_col_letter = get_column_letter(img_col_idx)

    # 列宽、行高大一点，尽量避免 Excel 再缩放图像
    ws.column_dimensions[img_col_letter].width = 20

    for idx, row in df.iterrows():
        smiles = row.get(smiles_col, None)
        atom_indices_raw = row.get(atom_idx_col, None)

        if pd.isna(smiles):
            continue

        if pd.isna(atom_indices_raw):
            atom_indices = []
        else:
            try:
                if isinstance(atom_indices_raw, str):
                    atom_indices = ast.literal_eval(atom_indices_raw)
                elif isinstance(atom_indices_raw, (list, tuple)):
                    atom_indices = list(atom_indices_raw)
                else:
                    atom_indices = [int(atom_indices_raw)]
            except Exception as e:
                print(f"Row {idx}: parse atom_index error: {atom_indices_raw} ({e})")
                continue

        try:
            img = highlight_num_atoms(
                smiles,
                atom_indices,
                size=img_size,
                internal_size=(400, 400),  # 内部画大图
            )
            img_path = os.path.join(img_dir, f"row_{idx}.png")
            img.save(img_path)

            xl_img = XLImage(img_path)
            # 保证插入到 Excel 时就是我们想要的大小（单位：像素）
            xl_img.width, xl_img.height = img_size

            excel_row = idx + 2
            cell = ws.cell(row=excel_row, column=img_col_idx)
            xl_img.anchor = cell.coordinate
            ws.add_image(xl_img)

            # 行高用点（pt），粗略按 0.75 * 像素
            ws.row_dimensions[excel_row].height = img_size[1] * 0.75

            if idx < 5:
                print(f"Row {idx}: inserted image {img_path}, atoms={atom_indices}")
        except Exception as e:
            print(f"Row {idx}: draw/save/insert error: {e}")
            continue

    wb.save(excel_out)
    print(f"\n完成：图片已嵌入到 {excel_out} 的 '{img_col}' 列。")
    print(f"PNG 图片保存在目录：{img_dir}/")


if __name__ == "__main__":
    head = '7_7_50k'
    attn = 'del'
    current_dir = os.getcwd()
    parent_dir = os.path.dirname(os.path.dirname(current_dir))
    network = "transformer"
    batch_size = 1
    epochs = 50
    learn_rate = 5e-05
    dropout = 0.3
    weight_decay = 0
    best_val_mcc = 0.72
    save_name = "network"
    result_file_name = f'data/result/statistics_supervision/uspto_yang/shield/ruizhen_product/{head}/result/{save_name}'
    result_folder = os.path.join(parent_dir, result_file_name)
    path_excel_atoms_input = (f'network_{network}_batch_{batch_size}_epochs_{epochs}_learn_rate_{learn_rate}_drop_{dropout}'
                              f'_weight_delay_{weight_decay}_head_{head}_attn_{attn}_best_val_mcc_{best_val_mcc}_atoms_index.xlsx')
    excel_in=os.path.join(result_folder, path_excel_atoms_input)
    path_excel_atoms_output = (f'network_{network}_batch_{batch_size}_epochs_{epochs}_learn_rate_{learn_rate}_drop_{dropout}'
                              f'_weight_delay_{weight_decay}_head_{head}_attn_{attn}_best_val_mcc_{best_val_mcc}_atoms_index_img.xlsx')
    excel_out = os.path.join(result_folder, path_excel_atoms_output)

    # excel_in = "ruizhen_reactantion_simple_rexgen_top6_merged_minus_index.xlsx"
    # excel_out = "ruizhen_reactantion_simple_rexgen_top6_merged_minus_index_img.xlsx"

    excel_embed_highlight_images(
        excel_in=excel_in,
        excel_out=excel_out,
        img_dir="atom_images",
        smiles_col="smiles",
        atom_idx_col="atom_index",
        img_col="atom_img",
        img_size=(120, 120),  # 可以改成 (100, 100) / (140, 140) 微调
    )

